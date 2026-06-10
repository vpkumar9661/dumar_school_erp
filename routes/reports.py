import io
from flask import Blueprint, render_template, request, send_file, make_response, flash, redirect, url_for, current_app
from flask_login import login_required
from db import db
from auth.decorators import permission_required
from datetime import datetime, date

reports_bp = Blueprint('reports', __name__)
CLASSES = ['Nursery','LKG','UKG'] + [str(i) for i in range(1,13)]
MONTHS  = ['January','February','March','April','May','June',
           'July','August','September','October','November','December']

def get_setting(key, default=''):
    try:
        cur = db.connection.cursor()
        cur.execute("SELECT setting_value FROM settings WHERE setting_key=%s",(key,))
        row = cur.fetchone(); cur.close()
        return row['setting_value'] if row else default
    except: return default

@reports_bp.route('/')
@login_required
@permission_required('manage_reports')
def index():
    return render_template('reports/index.html',
        classes=CLASSES, months=MONTHS,
        years=range(2020, datetime.now().year+2))

@reports_bp.route('/students/excel')
@login_required
@permission_required('manage_reports')
def students_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        cls    = request.args.get('class','')
        status = request.args.get('status','Active')
        cur = db.connection.cursor()
        q = "SELECT * FROM students WHERE 1=1"; p=[]
        if cls:    q += " AND class=%s";  p.append(cls)
        if status: q += " AND status=%s"; p.append(status)
        q += " ORDER BY class,section,roll_number"
        cur.execute(q,p); students = cur.fetchall(); cur.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Students"
        ws['A1'] = get_setting('school_name','VVM Dharampur')
        ws['A1'].font = Font(bold=True, size=14, color="00D4FF")
        ws.merge_cells('A1:N1')
        headers = ['Sr','Adm No','Name','Father','Class','Sec','Roll','Gender','DOB','Mobile','Monthly Fee','Transport','Status','Admission Date']
        fill = PatternFill("solid", fgColor="101828")
        for col,h in enumerate(headers,1):
            c = ws.cell(row=2,column=col,value=h)
            c.fill = fill; c.font = Font(bold=True,color="00D4FF")
            c.alignment = Alignment(horizontal='center')
        for i,s in enumerate(students,1):
            ws.append([i, s['admission_number'], s['student_name'], s['father_name'],
                       s['class'], s['section'], s['roll_number'], s['gender'],
                       str(s['date_of_birth'] or ''), s['mobile_number'],
                       float(s['monthly_fee']), s['transport_distance'],
                       s['status'], str(s['admission_date'] or '')])
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True,
                         download_name=f"Students_{cls or 'All'}_{date.today()}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        flash('openpyxl not installed.','warning')
        return redirect(url_for('reports.index'))

@reports_bp.route('/fees/excel')
@login_required
@permission_required('manage_reports')
def fees_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        month = request.args.get('month',''); year = request.args.get('year',''); cls = request.args.get('class','')
        cur = db.connection.cursor()
        q = """SELECT fp.*,s.student_name,s.admission_number,s.class,s.section
               FROM fee_payments fp JOIN students s ON s.id=fp.student_id WHERE 1=1"""
        p=[]
        if month: q += " AND fp.month=%s"; p.append(month)
        if year:  q += " AND fp.year=%s";  p.append(year)
        if cls:   q += " AND s.class=%s";  p.append(cls)
        q += " ORDER BY fp.created_at DESC"
        cur.execute(q,p); payments = cur.fetchall(); cur.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Fee Report"
        ws['A1'] = get_setting('school_name','VVM Dharampur')
        ws['A1'].font = Font(bold=True,size=14,color="00D4FF")
        ws.merge_cells('A1:L1')
        headers = ['Sr','Receipt','Student','Adm No','Class','Sec','Month','Year','Monthly Fee','Transport','Late Fine','Total']
        fill = PatternFill("solid", fgColor="101828")
        for col,h in enumerate(headers,1):
            c = ws.cell(row=2,column=col,value=h)
            c.fill = fill; c.font = Font(bold=True,color="00D4FF")
        for i,p2 in enumerate(payments,1):
            ws.append([i, p2['receipt_number'], p2['student_name'], p2['admission_number'],
                       p2['class'], p2['section'], p2['month'], p2['year'],
                       float(p2['monthly_fee']), float(p2['transport_fee']),
                       float(p2['late_fine']), float(p2['total_amount'])])
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, as_attachment=True,
                         download_name=f"FeeReport_{month or 'All'}_{year or 'All'}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        flash('openpyxl not installed.','warning')
        return redirect(url_for('reports.index'))
